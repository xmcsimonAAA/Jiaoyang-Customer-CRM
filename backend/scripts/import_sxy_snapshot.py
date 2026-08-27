#!/usr/bin/env python3
"""Import the 2026-08-21 SXY snapshot into the local SQLite CRM database.

The SXY master workbook is authoritative for TW customer numbers.  Ambiguous
legacy test rows are handled by an explicit, reviewable mapping below rather
than by a fuzzy name match.  Run without ``--apply`` first; the write mode makes
a SQLite backup and records every change in the audit log.
"""

from __future__ import annotations

import argparse
import json
import re
import shutil
import sqlite3
import unicodedata
from collections import Counter
from datetime import datetime
from decimal import Decimal, InvalidOperation
from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook
from opencc import OpenCC


ROOT = Path(__file__).resolve().parents[2]
DEFAULT_BASE = Path("/Users/simon/Desktop/骄阳市场数据系统/用所选项目新建的文件夹")
DEFAULT_MASTER = DEFAULT_BASE / "SXY 客户信息含1月15日后开户客户（动态进度）20260821 0630.xlsx"
DEFAULT_ASSET = DEFAULT_BASE / "SXY客户信息 20260821.xlsx"
DEFAULT_HOLDING = DEFAULT_BASE / "SXY SH 0821.xlsx"
DEFAULT_DB = ROOT / "customer_data.db"
DEFAULT_MAPPING = ROOT / "backups" / "sxy-legacy-mapping-20260821.json"
ACTOR_ID = "system-sxy-20260821"
ACTOR_NAME = "SXY 2026.08.21 数据同步"
SOURCE_DETAIL = "SXY 数据快照 · 2026.08.21"
ASSET_FIELD_LABEL = "券商账户资产（USD）"
HOLDING_DATE = "2026-08-21"
HOLDING_SECURITY = "二级市场持仓"

STATUS_MAP = {
    "未开户": "未启动",
    "未提交申请": "未启动",
    "提交中": "资料准备",
    "待初审资料": "资料准备",
    "待传住址证明": "资料准备",
    "待审住址证明": "资料准备",
    "处理中": "开户审核",
    "已開通": "已开户",
    "开户成功": "已开户",
    "开户失败": "开户失败",
    "駁回": "开户失败",
    "驳回": "开户失败",
    "开户资料驳回至客服": "开户失败",
    "开户资料驳回至客户": "开户失败",
}

def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description=__doc__)
    parser.add_argument("--master", type=Path, default=DEFAULT_MASTER)
    parser.add_argument("--asset", type=Path, default=DEFAULT_ASSET)
    parser.add_argument("--holding", type=Path, default=DEFAULT_HOLDING)
    parser.add_argument("--database", type=Path, default=DEFAULT_DB)
    parser.add_argument("--mapping", type=Path, default=DEFAULT_MAPPING, help="local ignored JSON mapping for legacy duplicate rows")
    parser.add_argument("--apply", action="store_true", help="write the validated import")
    return parser.parse_args()


def text(value: object) -> str:
    return "" if value is None else str(value).strip()


def normalized_code(value: object) -> str:
    return re.sub(r"\s+", "", text(value)).upper()


def tw_code(value: object) -> str:
    match = re.search(r"(?i)\b(TW\d+)\b", text(value))
    return match.group(1).upper() if match else ""


def normalized_name(value: object) -> str:
    converted = OpenCC("t2s").convert(unicodedata.normalize("NFKC", text(value)))
    converted = re.sub(r"^(微信名|wx|微信)\s*[:：]\s*", "", converted, flags=re.IGNORECASE)
    return re.sub(r"[\s\u3000]+", "", converted).casefold()


def decimal_value(value: object) -> Decimal:
    raw = text(value).replace(",", "").replace("$", "").replace("￥", "")
    if not raw or raw.upper() in {"#N/A", "N/A", "NA", "-", "/"}:
        return Decimal("0")
    try:
        return Decimal(raw)
    except InvalidOperation as exc:
        raise ValueError(f"无法识别金额或数量: {value!r}") from exc


def decimal_text(value: Decimal) -> str:
    return format(value, "f").rstrip("0").rstrip(".") or "0"


def load_master(path: Path) -> dict[str, dict[str, str]]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    records: dict[str, dict[str, str]] = {}
    for row in sheet.iter_rows(min_row=2, values_only=True):
        code = tw_code(row[2] if len(row) > 2 else None)
        name = text(row[0] if row else None)
        raw_status = text(row[1] if len(row) > 1 else None)
        if not code and not name and not raw_status:
            continue
        if not re.fullmatch(r"TW\d+", code):
            raise ValueError(f"开户状态表存在无效 TW 编号: {row[:3]}")
        if not name:
            raise ValueError(f"开户状态表存在空姓名: {code}")
        if code in records:
            raise ValueError(f"开户状态表存在重复 TW 编号: {code}")
        records[code] = {"code": code, "name": name, "raw_status": raw_status, "status": STATUS_MAP.get(raw_status, raw_status or "未启动")}
    return records


def load_asset(path: Path) -> tuple[dict[str, dict[str, object]], int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    records: dict[str, dict[str, object]] = {}
    invalid = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        code = tw_code(row[0] if row else None)
        if not code:
            invalid += 1
            continue
        if code in records:
            raise ValueError(f"资产表存在重复 TW 编号: {code}")
        records[code] = {"name": text(row[1] if len(row) > 1 else None), "asset": decimal_value(row[2] if len(row) > 2 else None)}
    return records, invalid


def load_holding(path: Path) -> tuple[dict[str, dict[str, object]], int]:
    workbook = load_workbook(path, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    records: dict[str, dict[str, object]] = {}
    invalid = 0
    for row in sheet.iter_rows(min_row=2, values_only=True):
        code = tw_code(row[0] if row else None)
        if not code:
            invalid += 1
            continue
        if code in records:
            raise ValueError(f"持仓表存在重复 TW 编号: {code}")
        records[code] = {"name": text(row[1] if len(row) > 1 else None), "quantity": decimal_value(row[2] if len(row) > 2 else None)}
    return records, invalid


def load_mapping(path: Path) -> tuple[dict[str, str], dict[str, str], set[str]]:
    if not path.is_file():
        raise FileNotFoundError(f"找不到本地旧客户映射文件（不会提交到 GitHub）: {path}")
    payload = json.loads(path.read_text(encoding="utf-8"))
    legacy_to_tw = {str(key): str(value) for key, value in payload.get("legacyToTw", {}).items()}
    duplicate_to_keep = {str(key): str(value) for key, value in payload.get("legacyDuplicateToKeep", {}).items()}
    new_unassigned = {str(value) for value in payload.get("newUnassignedCodes", [])}
    if not legacy_to_tw or not duplicate_to_keep or not new_unassigned:
        raise ValueError("本地旧客户映射文件缺少 legacyToTw、legacyDuplicateToKeep 或 newUnassignedCodes")
    return legacy_to_tw, duplicate_to_keep, new_unassigned


def now_iso() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def audit(conn: sqlite3.Connection, action: str, entity_type: str, entity_id: str, detail: dict[str, object]) -> None:
    conn.execute(
        """INSERT INTO audit_logs(id, actor_id, actor_name, action, entity_type, entity_id, detail_json, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
        (str(uuid4()), ACTOR_ID, ACTOR_NAME, action, entity_type, entity_id, json.dumps(detail, ensure_ascii=False), now_iso()),
    )


def backup_database(database: Path) -> Path:
    backup_dir = ROOT / "backups"
    backup_dir.mkdir(parents=True, exist_ok=True)
    backup_path = backup_dir / f"customer_data-before-sxy-{datetime.now().strftime('%Y%m%d-%H%M%S')}.db"
    source = sqlite3.connect(database)
    target = sqlite3.connect(backup_path)
    try:
        source.execute("PRAGMA wal_checkpoint(FULL)")
        source.backup(target)
    finally:
        target.close()
        source.close()
    backup_path.chmod(0o600)
    return backup_path


def ensure_asset_field(conn: sqlite3.Connection) -> str:
    row = conn.execute("SELECT id FROM customer_fields WHERE label=? AND active=1", (ASSET_FIELD_LABEL,)).fetchone()
    timestamp = now_iso()
    if row:
        return str(row[0])
    field_id = str(uuid4())
    field_key = f"custom_{field_id.replace('-', '')[:12]}"
    order = conn.execute("SELECT COALESCE(MAX(display_order), -1) + 1 FROM customer_fields").fetchone()[0]
    conn.execute(
        """INSERT INTO customer_fields(id, field_key, label, field_type, options_json, display_order, active,
        created_by, created_at, updated_by, updated_at) VALUES (?, ?, ?, 'number', '[]', ?, 1, ?, ?, ?, ?)""",
        (field_id, field_key, ASSET_FIELD_LABEL, order, ACTOR_ID, timestamp, ACTOR_ID, timestamp),
    )
    return field_id


def main() -> int:
    args = parse_args()
    database = args.database.expanduser().resolve()
    master = load_master(args.master.expanduser().resolve())
    assets, invalid_asset_rows = load_asset(args.asset.expanduser().resolve())
    holdings, invalid_holding_rows = load_holding(args.holding.expanduser().resolve())
    legacy_to_tw, legacy_duplicate_to_keep, new_unassigned_codes = load_mapping(args.mapping.expanduser().resolve())
    if not database.is_file():
        raise FileNotFoundError(database)
    unknown_asset_codes = sorted(set(assets) - set(master))
    unknown_holding_codes = sorted(set(holdings) - set(master))
    if unknown_asset_codes or unknown_holding_codes:
        raise ValueError(f"资产/持仓表存在主表没有的 TW 编号: asset={unknown_asset_codes}, holding={unknown_holding_codes}")

    conn = sqlite3.connect(database)
    conn.row_factory = sqlite3.Row
    conn.execute("PRAGMA foreign_keys = ON")
    try:
        integrity = conn.execute("PRAGMA integrity_check").fetchone()[0]
        if integrity != "ok":
            raise RuntimeError(f"SQLite 完整性检查失败: {integrity}")
        customers = {row["customer_code"]: row for row in conn.execute("SELECT * FROM customers WHERE archived_at IS NULL")}
        all_customer_codes = {row[0] for row in conn.execute("SELECT customer_code FROM customers")}
        current_tw = {code for code in customers if normalized_code(code).startswith("TW")}
        missing_before = set(master) - current_tw
        unexpected_current = current_tw - set(master)
        if unexpected_current:
            raise ValueError(f"当前数据库有主表之外的 TW 编号: {sorted(unexpected_current)}")
        if set(legacy_to_tw) - set(customers):
            raise ValueError(f"预设旧客户编号不存在: {sorted(set(legacy_to_tw) - set(customers))}")
        if set(legacy_duplicate_to_keep) - set(customers) or set(legacy_duplicate_to_keep.values()) - set(customers):
            raise ValueError("预设重复客户编号不存在")
        mapped_codes = set(legacy_to_tw.values())
        if len(mapped_codes) != len(legacy_to_tw):
            raise ValueError("预设映射中存在重复目标 TW 编号")
        if mapped_codes - missing_before:
            raise ValueError(f"预设映射目标已经存在，拒绝覆盖: {sorted(mapped_codes - missing_before)}")
        created_codes = missing_before - mapped_codes
        if created_codes != new_unassigned_codes:
            raise ValueError(f"缺失 TW 编号与预设新建清单不一致: {sorted(created_codes)}")
        if set(new_unassigned_codes) & all_customer_codes:
            raise ValueError("待新建 TW 编号已在数据库中占用")

        report = {
            "mode": "apply" if args.apply else "dry-run",
            "masterCodes": len(master),
            "currentActiveCustomers": len(customers),
            "alreadyMatchedCodes": len(current_tw),
            "legacyMapped": len(legacy_to_tw),
            "legacyArchivedDuplicates": len(legacy_duplicate_to_keep),
            "newUnassigned": len(new_unassigned_codes),
            "assetRows": len(assets),
            "assetInvalidRows": invalid_asset_rows,
            "holdingRows": len(holdings),
            "holdingInvalidRows": invalid_holding_rows,
            "assetCodesNotYetInCustomerPool": len(set(assets) - current_tw - mapped_codes),
            "holdingCodesNotYetInCustomerPool": len(set(holdings) - current_tw - mapped_codes),
        }
        print(json.dumps(report, ensure_ascii=False, indent=2))
        if not args.apply:
            print("Dry run completed. No local database was changed.")
            return 0

        backup_path = backup_database(database)
        timestamp = now_iso()
        conn.execute("BEGIN IMMEDIATE")
        by_code = {row["customer_code"]: row for row in conn.execute("SELECT * FROM customers WHERE archived_at IS NULL")}
        # Update every authoritative code's normalized account status. Existing
        # business ownership, Hongan advisor, and follow-up history remain intact.
        for code in sorted(current_tw):
            record = master[code]
            row = by_code[code]
            conn.execute("UPDATE customers SET name=?, account_status=?, updated_at=?, version=version+1 WHERE id=?", (record["name"], record["status"], timestamp, row["id"]))

        for old_code, target_code in legacy_to_tw.items():
            record = master[target_code]
            row = by_code[old_code]
            conn.execute("UPDATE customers SET customer_code=?, name=?, account_status=?, updated_at=?, version=version+1 WHERE id=?", (target_code, record["name"], record["status"], timestamp, row["id"]))
            audit(conn, "customer.sxy_code_migrated", "customer", row["id"], {"fromCode": old_code, "toCode": target_code, "rawStatus": record["raw_status"], "status": record["status"]})

        for old_code, keep_code in legacy_duplicate_to_keep.items():
            source = by_code[old_code]
            target = by_code[keep_code]
            conn.execute("UPDATE customers SET archived_at=?, merged_into_id=?, updated_at=?, version=version+1 WHERE id=?", (timestamp, target["id"], timestamp, source["id"]))
            conn.execute("INSERT INTO merge_events VALUES (?, ?, ?, ?, ?, ?, ?)", (str(uuid4()), source["id"], target["id"], "SXY 主表去重", ACTOR_ID, ACTOR_NAME, timestamp))
            audit(conn, "customer.sxy_duplicate_archived", "customer", source["id"], {"fromCode": old_code, "keptCode": keep_code, "reason": "主表只有一个对应 TW 编号"})

        for code in sorted(new_unassigned_codes):
            record = master[code]
            customer_id = str(uuid4())
            conn.execute(
                """INSERT INTO customers(id, customer_code, name, phone, email, wechat_nickname, company, source, source_detail,
                stage, priority, owner_id, owner_name, owner_team, notes, created_by, created_at, updated_at, account_status,
                account_broker, account_opened_at, broker_deposit_amount, capital_destination, intent_status, placement_status,
                target_batch_id, intent_amount, funded_amount, actual_amount, lost_reason, closed_at, hongan_advisor, source_advisor_label)
                VALUES (?, ?, ?, '', '', '', '', '历史存量', ?, '开户推进', '普通', 'unassigned', '待分配', '待分配池', '', ?, ?, ?, ?, '', NULL, 0, '', '未确认', '未进入', NULL, 0, 0, 0, '', NULL, '', '')""",
                (customer_id, code, record["name"], SOURCE_DETAIL, ACTOR_ID, timestamp, timestamp, record["status"]),
            )
            conn.execute("INSERT INTO assignments VALUES (?, ?, NULL, NULL, NULL, 'unassigned', '待分配', '待分配池', ?, ?, ?, ?)", (str(uuid4()), customer_id, "SXY 主表新增", ACTOR_ID, ACTOR_NAME, timestamp))
            audit(conn, "customer.sxy_created", "customer", customer_id, {"customerCode": code, "rawStatus": record["raw_status"], "status": record["status"]})

        # TW numbers are first-class identifiers.  customer_code remains the
        # display key for now, while this table enables future partial imports.
        active_by_code = {row["customer_code"]: row for row in conn.execute("SELECT * FROM customers WHERE archived_at IS NULL")}
        for code in sorted(master):
            customer = active_by_code.get(code)
            if not customer:
                raise RuntimeError(f"迁移后找不到客户: {code}")
            conn.execute("INSERT OR IGNORE INTO customer_identifiers(customer_id, kind, normalized_value, display_value, created_at) VALUES (?, 'tw', ?, ?, ?)", (customer["id"], code, code, timestamp))

        asset_field_id = ensure_asset_field(conn)
        for code, asset in assets.items():
            customer = active_by_code.get(code)
            if not customer:
                continue
            value = decimal_text(asset["asset"])
            conn.execute(
                """INSERT INTO customer_field_values(customer_id, field_id, value_text, updated_by, updated_at) VALUES (?, ?, ?, ?, ?)
                ON CONFLICT(customer_id, field_id) DO UPDATE SET value_text=excluded.value_text, updated_by=excluded.updated_by, updated_at=excluded.updated_at""",
                (customer["id"], asset_field_id, value, ACTOR_ID, timestamp),
            )

        for code, holding in holdings.items():
            customer = active_by_code.get(code)
            if not customer:
                continue
            quantity = float(holding["quantity"])
            conn.execute(
                """INSERT INTO customer_holding_snapshots(id, customer_id, snapshot_date, security_name, quantity, market_value, source_label, created_by, created_at)
                VALUES (?, ?, ?, ?, ?, 0, ?, ?, ?)
                ON CONFLICT(customer_id, snapshot_date, security_name) DO UPDATE SET quantity=excluded.quantity, source_label=excluded.source_label, created_by=excluded.created_by, created_at=excluded.created_at""",
                (str(uuid4()), customer["id"], HOLDING_DATE, HOLDING_SECURITY, quantity, "SXY SH 2026.08.21", ACTOR_ID, timestamp),
            )

        audit(conn, "migration.sxy_snapshot_imported", "migration", "sxy-20260821", {**report, "backup": str(backup_path), "assetField": ASSET_FIELD_LABEL, "holdingDate": HOLDING_DATE})
        conn.commit()
        print(json.dumps({"migration": "completed", "backup": str(backup_path), "activeCustomersAfter": len(master), "twIdentifiersAfter": conn.execute("SELECT COUNT(*) FROM customer_identifiers WHERE kind='tw'").fetchone()[0]}, ensure_ascii=False, indent=2))
        return 0
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    raise SystemExit(main())
