from __future__ import annotations

import json
import re
import shutil
import sqlite3
import unicodedata
from collections import defaultdict
from datetime import datetime
from pathlib import Path
from uuid import uuid4

from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parents[2]
DB_PATH = ROOT / "customer_data.db"
SOURCE_PATH = Path(
    "/Users/simon/Desktop/骄阳市场数据系统/用所选项目新建的文件夹/"
    "SXY 客户信息含1月15日后开户客户（动态进度）20260821 0630.xlsx"
)
BACKUP_PATH = ROOT / "backups" / "customer_data-before-sxy-tw-migration-20260824-v2.db"
ACTOR_ID = "system-migration"
ACTOR_NAME = "SXY 编号迁移"
UNASSIGNED_OWNER_ID = "unassigned"
UNASSIGNED_OWNER_NAME = "待分配"
UNASSIGNED_OWNER_TEAM = "待分配池"
SOURCE_DETAIL = "SXY开户状态表 · 2026.08.21"

STATUS_MAP = {
    "未开户": "未启动",
    "未提交申请": "未启动",
    "提交中": "资料准备",
    "待初审资料": "资料准备",
    "待传住址证明": "资料准备",
    "待审住址证明": "资料准备",
    "处理中": "开户审核",
    "已開通": "已开户",
    "开户成功": "已开户",
    "开户失败": "开户失败",
    "駁回": "开户失败",
    "驳回": "开户失败",
    "开户资料驳回至客服": "开户失败",
    "开户资料驳回至客户": "开户失败",
}


def text(value: object) -> str:
    return "" if value is None else str(value).strip()


def normalized_name(value: object) -> str:
    value = unicodedata.normalize("NFKC", text(value))
    value = re.sub(r"^(微信名|wx|微信)\s*[:：]\s*", "", value, flags=re.IGNORECASE)
    return re.sub(r"[\s\u3000]+", "", value).casefold()


def normalized_code(value: object) -> str:
    return text(value).replace(" ", "").upper()


def load_master() -> list[dict[str, str]]:
    workbook = load_workbook(SOURCE_PATH, read_only=True, data_only=True)
    sheet = workbook.worksheets[0]
    records: list[dict[str, str]] = []
    seen: set[str] = set()
    for row in sheet.iter_rows(min_row=2, values_only=True):
        name, status, raw_code = (row + (None, None, None))[:3]
        record = {
            "name": text(name),
            "raw_status": text(status),
            "status": STATUS_MAP.get(text(status), text(status) or "未启动"),
            "code": normalized_code(raw_code),
        }
        if not any(record.values()):
            continue
        if not re.fullmatch(r"TW\d+", record["code"]):
            raise RuntimeError(f"开户表存在无效客户编号: {record}")
        if not record["name"]:
            raise RuntimeError(f"开户表存在空客户姓名: {record}")
        if record["code"] in seen:
            raise RuntimeError(f"开户表存在重复客户编号: {record['code']}")
        seen.add(record["code"])
        records.append(record)
    return records


def now_iso() -> str:
    return datetime.now().astimezone().isoformat(timespec="seconds")


def audit(conn: sqlite3.Connection, action: str, entity_type: str, entity_id: str, detail: dict) -> None:
    conn.execute(
        """INSERT INTO audit_logs(id, actor_id, actor_name, action, entity_type, entity_id, detail_json, created_at)
        VALUES (?, ?, ?, ?, ?, ?, ?, ?)""",
        (str(uuid4()), ACTOR_ID, ACTOR_NAME, action, entity_type, entity_id, json.dumps(detail, ensure_ascii=False), now_iso()),
    )


def main() -> None:
    if not SOURCE_PATH.is_file():
        raise FileNotFoundError(SOURCE_PATH)
    if BACKUP_PATH.exists():
        raise FileExistsError(f"备份文件已存在，拒绝覆盖: {BACKUP_PATH}")

    master = load_master()
    by_name: dict[str, list[dict[str, str]]] = defaultdict(list)
    for record in master:
        by_name[normalized_name(record["name"])].append(record)

    BACKUP_PATH.parent.mkdir(parents=True, exist_ok=True)
    source = sqlite3.connect(DB_PATH)
    try:
        backup = sqlite3.connect(BACKUP_PATH)
        try:
            source.backup(backup)
        finally:
            backup.close()
    finally:
        source.close()

    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    timestamp = now_iso()
    try:
        conn.execute("PRAGMA foreign_keys = ON")
        conn.execute("BEGIN IMMEDIATE")
        customers = conn.execute("SELECT * FROM customers WHERE archived_at IS NULL ORDER BY customer_code").fetchall()
        current_codes = {normalized_code(row["customer_code"]) for row in customers}
        all_codes = {normalized_code(row["customer_code"]) for row in conn.execute("SELECT customer_code FROM customers").fetchall()}

        provisional_matches: list[tuple[sqlite3.Row, dict[str, str]]] = []
        ambiguous: list[dict] = []
        unmatched: list[sqlite3.Row] = []
        for customer in customers:
            candidates: dict[str, dict[str, str]] = {}
            for field in ("name", "wechat_nickname"):
                key = normalized_name(customer[field])
                if not key:
                    continue
                for record in by_name.get(key, []):
                    candidates[record["code"]] = record
            if len(candidates) == 1:
                provisional_matches.append((customer, next(iter(candidates.values()))))
            elif len(candidates) > 1:
                ambiguous.append({"customerCode": customer["customer_code"], "name": customer["name"], "wechatNickname": customer["wechat_nickname"], "candidateCodes": sorted(candidates)})
            else:
                unmatched.append(customer)

        mapped_by_code: dict[str, list[sqlite3.Row]] = defaultdict(list)
        for customer, record in provisional_matches:
            mapped_by_code[record["code"]].append(customer)
        duplicate_code_groups = {
            code: rows for code, rows in mapped_by_code.items() if len(rows) > 1
        }
        duplicate_codes = set(duplicate_code_groups)
        matches = [
            (customer, record)
            for customer, record in provisional_matches
            if record["code"] not in duplicate_codes
        ]
        for code, rows in duplicate_code_groups.items():
            ambiguous.append({
                "customerCode": None,
                "name": rows[0]["name"],
                "wechatNickname": rows[0]["wechat_nickname"],
                "candidateCodes": [code],
                "reason": "同一 TW 编号对应多条旧测试记录，需人工合并",
                "existingCustomerCodes": [row["customer_code"] for row in rows],
            })

        matched_codes = {record["code"] for _, record in matches}
        ambiguous_codes = {code for item in ambiguous for code in item["candidateCodes"]}
        master_by_code = {record["code"]: record for record in master}
        new_records = [record for record in master if record["code"] not in matched_codes and record["code"] not in ambiguous_codes]
        collisions = sorted(record["code"] for record in new_records if record["code"] in all_codes)
        if collisions:
            raise RuntimeError(f"待新增 TW 编号与数据库已有编号冲突: {collisions}")

        renamed = 0
        status_updated = 0
        name_filled = 0
        for customer, record in matches:
            changes: list[str] = []
            if customer["customer_code"] != record["code"]:
                changes.append("customer_code")
            new_name = customer["name"] or record["name"]
            if not customer["name"] and new_name:
                changes.append("name")
            if customer["account_status"] != record["status"]:
                changes.append("account_status")
            conn.execute(
                "UPDATE customers SET customer_code=?, name=?, account_status=?, updated_at=?, version=version+1 WHERE id=?",
                (record["code"], new_name, record["status"], timestamp, customer["id"]),
            )
            if "customer_code" in changes:
                renamed += 1
            if "name" in changes:
                name_filled += 1
            if "account_status" in changes:
                status_updated += 1
            audit(conn, "customer.sxy_code_migrated", "customer", customer["id"], {"fromCode": customer["customer_code"], "toCode": record["code"], "rawStatus": record["raw_status"], "status": record["status"], "nameFilled": "name" in changes})

        created = 0
        for record in new_records:
            customer_id = str(uuid4())
            conn.execute(
                """INSERT INTO customers(
                    id, customer_code, name, phone, email, company, source, source_detail, stage, priority,
                    owner_id, owner_name, owner_team, notes, created_by, created_at, updated_at,
                    account_status, account_broker, account_opened_at, broker_deposit_amount,
                    capital_destination, intent_status, placement_status, target_batch_id,
                    intent_amount, funded_amount, actual_amount, lost_reason, closed_at,
                    hongan_advisor, source_advisor_label
                ) VALUES (?, ?, ?, '', '', '', '历史存量', ?, '开户推进', '普通', ?, ?, ?, '', ?, ?, ?, ?, '', NULL, 0, '', '未确认', '未进入', NULL, 0, 0, 0, '', NULL, '', '')""",
                (customer_id, record["code"], record["name"], SOURCE_DETAIL, UNASSIGNED_OWNER_ID, UNASSIGNED_OWNER_NAME, UNASSIGNED_OWNER_TEAM, ACTOR_ID, timestamp, timestamp, record["status"]),
            )
            conn.execute(
                "INSERT INTO assignments VALUES (?, ?, NULL, NULL, NULL, ?, ?, ?, ?, ?, ?, ?)",
                (str(uuid4()), customer_id, UNASSIGNED_OWNER_ID, UNASSIGNED_OWNER_NAME, UNASSIGNED_OWNER_TEAM, "SXY 开户表导入", ACTOR_ID, ACTOR_NAME, timestamp),
            )
            audit(conn, "customer.sxy_created", "customer", customer_id, {"customerCode": record["code"], "rawStatus": record["raw_status"], "status": record["status"]})
            created += 1

        archived = 0
        for customer in unmatched:
            conn.execute("UPDATE customers SET archived_at=?, updated_at=?, version=version+1 WHERE id=? AND archived_at IS NULL", (timestamp, timestamp, customer["id"]))
            audit(conn, "customer.sxy_archived", "customer", customer["id"], {"customerCode": customer["customer_code"], "reason": "最新 SXY 开户表中没有对应唯一编号"})
            archived += 1

        summary = {
            "masterRows": len(master),
            "matchedAndRenamed": renamed,
            "statusUpdated": status_updated,
            "namesFilled": name_filled,
            "created": created,
            "archived": archived,
            "ambiguous": len(ambiguous),
            "ambiguousCandidateCodes": len(ambiguous_codes),
            "duplicateMatchedCodes": len(duplicate_code_groups),
            "duplicateMatchedRows": sum(len(rows) for rows in duplicate_code_groups.values()),
            "ambiguousRows": ambiguous,
            "backup": str(BACKUP_PATH),
        }
        audit(conn, "migration.sxy_customer_codes", "migration", "sxy-20260824", summary)
        conn.commit()
        print(json.dumps(summary, ensure_ascii=False, indent=2))
    except Exception:
        conn.rollback()
        raise
    finally:
        conn.close()


if __name__ == "__main__":
    main()
